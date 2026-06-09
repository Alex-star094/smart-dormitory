"""
数据导出 API - 支持 Excel (xlsx) 格式导出
"""
from datetime import datetime
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from api.auth import get_current_user, require_admin
from utils.db import get_db
from models.user import User
from models.access import AccessRecord
from models.energy_consumption import EnergyConsumption
from models.repair_record import RepairRecord
from models.visitor import Visitor

router = APIRouter()

# 通用样式
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4361EE", end_color="4361EE", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
CELL_ALIGN = Alignment(vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _style_sheet(ws, headers: list, col_widths: list = None):
    """给工作表应用统一样式"""
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"


from urllib.parse import quote

def _make_response(wb: Workbook, filename: str) -> StreamingResponse:
    """生成 StreamingResponse（正确处理中文文件名）"""
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    encoded_filename = quote(filename)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
        },
    )


# ==================== 通行记录导出 ====================
@router.get("/access-records", summary="导出通行记录为 Excel")
def export_access_records(
    dormitory: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    start_date: Optional[str] = Query(None, description="起始日期 YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="结束日期 YYYY-MM-DD"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """导出通行记录（学生只能导出自己的，管理员可筛选导出全部）"""
    query = db.query(AccessRecord)

    if current_user.role == "student":
        query = query.filter(AccessRecord.user_id == current_user.id)
    else:
        if dormitory:
            query = query.filter(AccessRecord.dormitory == dormitory)
        if status:
            query = query.filter(AccessRecord.status == status.lower())

    if start_date:
        query = query.filter(AccessRecord.access_time >= datetime.strptime(start_date, "%Y-%m-%d"))
    if end_date:
        query = query.filter(AccessRecord.access_time <= datetime.strptime(end_date, "%Y-%m-%d") + " 23:59:59")

    records = query.order_by(AccessRecord.access_time.desc()).limit(5000).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "通行记录"
    headers = ["ID", "姓名", "宿舍", "状态", "通行方式", "通行时间", "告警", "备注"]
    _style_sheet(ws, headers, [6, 12, 10, 10, 10, 22, 8, 30])

    status_map = {"allowed": "允许", "denied": "拒绝", "error": "错误"}
    for row_idx, r in enumerate(records, 2):
        ws.cell(row=row_idx, column=1, value=r.id).border = THIN_BORDER
        ws.cell(row=row_idx, column=2, value=r.username).border = THIN_BORDER
        ws.cell(row=row_idx, column=3, value=r.dormitory).border = THIN_BORDER
        ws.cell(row=row_idx, column=4, value=status_map.get(r.status, r.status)).border = THIN_BORDER
        ws.cell(row=row_idx, column=5, value=r.access_type).border = THIN_BORDER
        ws.cell(row=row_idx, column=6, value=r.access_time.strftime("%Y-%m-%d %H:%M:%S") if r.access_time else "").border = THIN_BORDER
        ws.cell(row=row_idx, column=7, value="是" if r.alarm else "否").border = THIN_BORDER
        ws.cell(row=row_idx, column=8, value=r.notes or "").border = THIN_BORDER
        for c in range(1, 9):
            ws.cell(row=row_idx, column=c).alignment = CELL_ALIGN

    return _make_response(wb, f"access_records_{datetime.now().strftime('%Y%m%d')}.xlsx")


# ==================== 能耗导出 ====================
@router.get("/energy", summary="导出能耗记录为 Excel")
def export_energy(
    dormitory: Optional[str] = Query(None),
    month: Optional[str] = Query(None, description="月份 YYYY-MM"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """导出能耗记录（学生只能导出自己宿舍）"""
    query = db.query(EnergyConsumption)

    if current_user.role == "student":
        if not current_user.dormitory:
            raise HTTPException(400, "请先绑定宿舍")
        query = query.filter(EnergyConsumption.dormitory == current_user.dormitory)
    elif dormitory:
        query = query.filter(EnergyConsumption.dormitory == dormitory)

    if month:
        query = query.filter(EnergyConsumption.month == month)

    records = query.order_by(EnergyConsumption.month.desc(), EnergyConsumption.dormitory).limit(5000).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "能耗记录"
    headers = ["ID", "宿舍", "能源类型", "消耗量", "单位", "费用(元)", "月份", "告警", "告警原因", "记录时间"]
    _style_sheet(ws, headers, [6, 10, 10, 10, 8, 10, 10, 8, 30, 22])

    type_map = {"electricity": "电力", "water": "用水", "gas": "燃气"}
    for row_idx, r in enumerate(records, 2):
        vals = [r.id, r.dormitory, type_map.get(r.energy_type, r.energy_type),
                r.consumption, r.unit, r.cost, r.month, "是" if r.alarm else "否",
                r.alarm_reason or "", r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else ""]
        for col_idx, v in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.border = THIN_BORDER; cell.alignment = CELL_ALIGN

    return _make_response(wb, f"energy_records_{datetime.now().strftime('%Y%m%d')}.xlsx")


# ==================== 维修记录导出（管理员） ====================
@router.get("/repairs", summary="导出维修记录为 Excel（管理员）")
def export_repairs(
    status: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """管理员导出维修记录"""
    query = db.query(RepairRecord)
    if status:
        query = query.filter(RepairRecord.status == status.lower())
    records = query.order_by(RepairRecord.created_at.desc()).limit(5000).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "维修记录"
    headers = ["ID", "宿舍", "标题", "类别", "优先级", "状态", "位置", "电话", "维修结果", "费用", "创建时间"]
    _style_sheet(ws, headers, [6, 10, 20, 10, 8, 10, 20, 15, 25, 10, 22])

    status_map = {"pending": "待处理", "processing": "处理中", "completed": "已完成", "cancelled": "已取消"}
    for row_idx, r in enumerate(records, 2):
        vals = [r.id, r.dormitory, r.title, r.category, r.priority,
                status_map.get(r.status, r.status), r.location, r.contact_phone,
                r.repair_result or "", r.cost or 0,
                r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else ""]
        for col_idx, v in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.border = THIN_BORDER; cell.alignment = CELL_ALIGN

    return _make_response(wb, f"repair_records_{datetime.now().strftime('%Y%m%d')}.xlsx")
